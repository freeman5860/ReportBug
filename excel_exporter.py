# -*- coding: utf-8 -*-
"""
Excel导出模块
"""
from datetime import datetime
from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models import Issue
from config import EXPORT_DIR


class ExcelExporter:
    """Excel导出器"""
    
    # 表头定义
    HEADERS = [
        "序号",
        "问题描述",
        "反馈人",
        "问题创建时间",
        "状态",
        "优先级",
        "问题分类",
        "备注"
    ]
    
    def __init__(self):
        self.export_dir = EXPORT_DIR
    
    def export_issues(self, issues: List[Issue], filename: str = None) -> Path:
        """导出问题到Excel文件"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"问题跟进记录_{timestamp}.xlsx"
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = self.export_dir / filename
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "问题跟进记录"
        
        # 设置列宽
        column_widths = [8, 40, 12, 20, 12, 10, 12, 30]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[self._get_column_letter(i)].width = width
        
        # 写入表头
        self._write_header(ws)
        
        # 写入数据
        self._write_data(ws, issues)
        
        # 保存文件
        wb.save(filepath)
        
        return filepath
    
    def _write_header(self, ws):
        """写入表头"""
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = self._get_border()
        
        for col, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        ws.row_dimensions[1].height = 25
    
    def _write_data(self, ws, issues: List[Issue]):
        """写入数据"""
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = self._get_border()
        
        for row_idx, issue in enumerate(issues, start=2):
            # 序号
            cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = border
            
            # 问题描述
            cell = ws.cell(row=row_idx, column=2, value=issue.description)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # 反馈人
            cell = ws.cell(row=row_idx, column=3, value=issue.reporter)
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = border
            
            # 问题创建时间
            cell = ws.cell(row=row_idx, column=4, value=issue.created_at)
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = border
            
            # 状态
            cell = ws.cell(row=row_idx, column=5, value=issue.status)
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = border
            if issue.status == "已解决":
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif issue.status == "处理中":
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            
            # 优先级
            cell = ws.cell(row=row_idx, column=6, value=issue.priority)
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = border
            if issue.priority == "高":
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif issue.priority == "中":
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            
            # 问题分类
            cell = ws.cell(row=row_idx, column=7, value=issue.category)
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = border
            
            # 备注
            cell = ws.cell(row=row_idx, column=8, value=issue.remarks)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            ws.row_dimensions[row_idx].height = 30
    
    @staticmethod
    def _get_border():
        """获取边框样式"""
        side = Side(style='thin', color='000000')
        return Border(left=side, right=side, top=side, bottom=side)
    
    @staticmethod
    def _get_column_letter(col_idx):
        """获取列字母"""
        result = ""
        while col_idx > 0:
            col_idx -= 1
            result = chr(col_idx % 26 + 65) + result
            col_idx //= 26
        return result


# 全局导出器实例
exporter = ExcelExporter()
